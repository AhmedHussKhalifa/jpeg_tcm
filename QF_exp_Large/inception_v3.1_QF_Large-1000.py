# -*- coding: utf-8 -*-
# Inception image recognition attempt
import tensorflow as tf
import numpy as np
import re
import os
import time
# import sys
import logging

import openpyxl

# Change class ID to human readable label
class NodeLookup(object):
    def __init__(self):
        # Path for imagenet_2012_challenge_label_map_proto.pbtxt and imagenet_synset_to_human_label_map.txt
        label_lookup_path = '../../../5a.Google-Inception/Attempts/inception_model/imagenet_2012_challenge_label_map_proto.pbtxt'   
        uid_lookup_path = '../../../5a.Google-Inception/Attempts/inception_model/imagenet_synset_to_human_label_map.txt'
        self.node_lookup = self.load(label_lookup_path, uid_lookup_path)

    def load(self, label_lookup_path, uid_lookup_path):
        # Load UID look up file
        proto_as_ascii_lines = tf.gfile.GFile(uid_lookup_path).readlines()
        uid_to_human = {}
        # Read Data line by line
        for line in proto_as_ascii_lines :
            # Remove \n
            line=line.strip('\n')
            # Split by \t
            parsed_items = line.split('\t')
            # Get the label number
            uid = parsed_items[0]
            # Get the human readable label name
            human_string = parsed_items[1]
            # Save the label
            uid_to_human[uid] = human_string

        # Load label look up file
        proto_as_ascii = tf.gfile.GFile(label_lookup_path).readlines()
        node_id_to_uid = {}
        for line in proto_as_ascii:
            if line.startswith('  target_class:'):
                # Get number 1-1000
                target_class = int(line.split(': ')[1])
            if line.startswith('  target_class_string:'):
                # Get UID n********
                target_class_string = line.split(': ')[1]
                # Save the relationship between number (1-1000) and the uid (n********)
                node_id_to_uid[target_class] = target_class_string[1:-2]

        # Build the relationship between number (1-1000) and the label
        node_id_to_name = {}
        for key, val in node_id_to_uid.items():
            # Get the label
            name = uid_to_human[val]
            # Build the relationship between number (1-1000) and the label
            node_id_to_name[key] = name
        return node_id_to_name

    # Input 1-1000 Return corresponding label
    def id_to_string(self, node_id):
        if node_id not in self.node_lookup:
            return ''
        return self.node_lookup[node_id]

# Read pre-trained Inception-V3 to create graph
def create_graph():
  # the class that's been created from the textual definition in graph.proto
  # classify_image_graph_def.pb
	with tf.gfile.FastGFile('../../../5a.Google-Inception/Attempts/inception_model/classify_image_graph_def.pb', 'rb') as f:
		graph_def = tf.GraphDef()
		graph_def.ParseFromString(f.read())
		tf.import_graph_def(graph_def, name='')

os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3' # Hide the warning information from Tensorflow - annoying...

sess=tf.Session()
#create graph
create_graph()

# Open the excel sheet
# record = xlrd.open_workbook('Accuracy Record ILSVRC2012 QF.xls', formatting_info=True)
# recordnew = copy(record)
# sheet = recordnew.get_sheet(0)
# style = XFStyle()
# style.num_format_str = 'general'

wb = openpyxl.load_workbook('Accuracy Record ILSVRC2012 QF 10003.xlsx', read_only = False, keep_vba = True)
sheet=wb['Sheet1']

###### CHANGE FOR DIFFERENT CASE ######
# Group Test
rootdir = 'C:/Users/y77jiang/Downloads/ILSVRC2012/7/'
file_name_initial = 'ILSVRC2012_val_'
# file_suffix = '.jpg';
file_suffix = '.JPEG';
list = os.listdir(rootdir) # List all file in the folder
start_ID = 1;
End_ID = 1001;

######
# Place of identify 
counter = 0

# Image Range
for ID in range(start_ID,End_ID):
    file_name_number = str(ID).zfill(8) 
    file_name = file_name_initial + file_name_number

    # Quality Factor Range
    for i in range(20,11,-10):
        path =  rootdir + file_name + '-QF-'+ str(i) + file_suffix #os.path.join(rootdir,list[i])
        # print(path)
        if os.path.isfile(path):
            # Read the image file in Tf lib |decode style of the photo: 'rb' = non-UTF-8; 'r' = UTF-8
            image_data = tf.gfile.FastGFile(path, 'rb').read()

            #Inception-v3: last layer is output as softmax
            softmax_tensor= sess.graph.get_tensor_by_name('softmax:0')
            #Input the image, obtain the softmax prob value（one shape=(1,1008) vector）
            predictions = sess.run(softmax_tensor,{'DecodeJpeg/contents:0': image_data})
            #(1,1008)->(1008,)
            predictions = np.squeeze(predictions)

            # ID --> English string label.
            node_lookup = NodeLookup()

            ###### Staring of Comment Place ######
            # Others 
            # sheet_r = record.sheets()[0]
            # Label = wb['A']

            for j in range(0,1008):
                sheet.cell(row=(ID-1)*1009+2+j+1, column=4).value = predictions[j].item()
            top = predictions.argsort()[-2000:][::-1]

            counter = counter + 1
            rank = 0
            # location = -i/2+57
            first_5 = 0
            first_1 = 0
            found = False
            for node_id in top:
                rank = rank + 1           
                human_string = node_lookup.id_to_string(node_id)
                score = predictions[node_id]

                # print('%s (score = %.5f)' % (human_string, score))
                if(sheet["B"+str((ID-1)*1009+2)].value == human_string):
                    # -- Original -- 3 4
                    print('%d-%d: %s (score = %.5f)' % (ID,i,human_string, score))
                    sheet.cell(row=(ID-1)*1009+2, column=3).value = rank
                    sheet.cell(row=(ID-1)*1009+2, column=4).value = score.item()
                    # sheet.write(ID, 7, rank, style)
                    # sheet.write(ID, 8, score.item(), style)
                    # if(rank <= 5):
                    # 	sheet.write(ID, 9, 1, style)
                    # else:
                    # 	sheet.write(ID, 9, 0, style)

sess.close()

print('Done...')
# recordnew.save('Accuracy Record ILSVRC2012 QF-100.xls')
wb.save('Accuracy Record ILSVRC2012 QF 1000-20.xlsm')